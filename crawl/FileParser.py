from openpyxl import load_workbook

class FileParser(object):

    def parser(self,root_file):
        if root_file is None:
            return

        new_urls=self._get_new_urls(root_file)
        return new_urls

    def _get_new_urls(self,root_file):
        new_urls=set()
        #读取Excel文件
        urls= self.read_all_urls(root_file)
        for url in urls:
            new_urls.add(url)
        return new_urls


    def read_all_urls(self,filePath):

        wb = load_workbook(filePath)
        name=wb.sheetnames[0]
        ws=wb[name]

        urls=[]
        for row in ws.iter_rows(min_row=2,min_col=2):
            for cell in row:
                if(cell.value and cell.value.startswith('http')):
                    urls.append(cell.value)
                    
        wb.close()
        return urls
