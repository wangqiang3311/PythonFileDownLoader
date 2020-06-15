import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class DataOutput(object):

    def store_data(self,data,origin_path,new_url):
        if data is None:
            return
        try:
            filename=new_url.split("/")[-1]

            filepath=os.path.join("crawl/outputfiles",filename)

            with open(filepath, "wb") as code:
                code.write(data)

            self.mark_result(origin_path,new_url,True)

        except Exception as err:
            print("store data faild:"+str(err))
            self.mark_result(origin_path,new_url,False)

    def mark_result(self,origin_path,new_url,isok):
        #标记原始文件
        bgcolor='FF0000'
        if isok:
            bgcolor='AACF91'
            print("ok")
        
        self.mark_cell(origin_path,new_url,bgcolor)

    def mark_cell(self,origin_path,new_url,bgcolor):
    
        wb = load_workbook(origin_path)
        name=wb.sheetnames[0]
        ws=wb[name]

        for row in ws.iter_rows(min_row=2,min_col=2):
            for cell in row:
                if cell.value==new_url:
                    fill = PatternFill(fill_type = "solid", start_color=bgcolor,end_color=bgcolor)
                    cell.fill=fill
                    #cell.fill=fills.GradientFill(stop=['FF0000', '0000FF'])  #填充渐变
                    
        wb.save(origin_path)

